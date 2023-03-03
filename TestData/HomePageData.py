import openpyxl
class HomePageData:

    # test_homepage_data = [{"name":"vinayak kamati","email":"vinaykamate4757@gmail.com","password":"vinay9000","gender":"Male","DOB":"1998-10-12"},
    #                       {"name":"anshika shetty","email":"anshika@gmail.com","password":"anshika1234","gender":"Female","DOB":"2000-12-12"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict={}
        book = openpyxl.load_workbook("C:\\Users\\HP\\PycharmProjects\\pythonProject2\\Book1.xlsx")
        sheet = book.active
        for i in range(1,sheet.max_row+1):
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2,sheet.max_column+1):
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return[Dict]
